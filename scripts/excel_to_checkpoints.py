#!/usr/bin/env python3
"""Compile ops Excel (gi_authoring_*.xlsx) → checkpoints + checkspecs.

Production contract for Applies when:
  - blank → always
  - @C1 / @C1 AND @C2 / @C1 OR @C2 / NOT @C1 → resolve Condition rows
  - anything else → hard fail (no fuzzy free-text)
"""

from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../src")))
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

from openpyxl import load_workbook  # noqa: E402

from build_gi_authoring_excel import HEADERS  # noqa: E402
from check_block import compile_block  # noqa: E402
from checkspec import resolve_specs  # noqa: E402
from harness_rules import (  # noqa: E402
    _parse_linked_when,
    _slug,
    block_to_check_block,
    block_to_rule,
    is_linked_when,
    is_structured_when,
)
from obligation import save_checkspecs  # noqa: E402
from rules_to_checkpoints import rules_to_checkpoints  # noqa: E402

ACTION_MAP = {
    "Rule type not yet added": "manual review",
    "Must be filled in": "is present",
    "Must be in English": "is in English",
    "Must equal": "equals",
    "Must be one of": "is one of",
    "Must contain text": "contains phrase",
    "Must not contain text": "term must not appear",
    "Word must not appear": "term must not appear",
    "Number greater than": "number greater than",
    "Number less than": "number less than",
    "Defect type includes": "defect type includes",
    "Compare number": "number greater than",  # legacy label
    "Matches pattern": "matches pattern",
    "At least N": "at least N photos",
    "At most N": "at most N photos",
    "Ratio at least": "ratio at least",
    "File name matches": "filename matches",
    "AI document check": "LLM quote then match",
    "AI yes/no (flag if false)": "LLM yes/no",
    "AI photo check (flag if false)": "needs vision",
    "Manual review": "manual review",
}


def _validate_applies_when(row_id: str, when: str) -> None:
    text = (when or "").strip()
    if not text:
        return
    # Dropdown may show "(always — blank)" — treat as blank.
    if text.lower().startswith("(always"):
        return
    if is_linked_when(text):
        if not re.fullmatch(r"(?i)(NOT\s+)?@[\w.]+(\s+(AND|OR)\s+(NOT\s+)?@[\w.]+)*", text):
            raise SystemExit(
                f"[excel] {row_id}: Applies when {when!r} looks like a Condition link but "
                f"is malformed. Use @C1 or @C1 AND @C2 (dropdown)."
            )
        return
    if is_structured_when(text):
        # Advanced escape hatch (engineers / markdown parity). Prefer @C links for ops.
        print(
            f"[excel] NOTE: {row_id} uses structured Applies when {when!r}. "
            f"Prefer a Condition row + @C link for ops-authored files.",
            file=sys.stderr,
        )
        return
    raise SystemExit(
        f"[excel] {row_id}: REJECTED Applies when {when!r}.\n"
        f"  Ops must use blank (always) or @C1 / @C1 AND @C2.\n"
        f"  Build a Condition row (Look at + Rule type + Value), then link it with @C1.\n"
        f"  Free prose like 'when defects are found' is not allowed."
    )


def load_excel_blocks(xlsx_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(xlsx_path)
    ws = wb["Checks"] if "Checks" in wb.sheetnames else wb.active
    blocks: list[dict[str, str]] = []
    for row_idx in range(4, ws.max_row + 1):
        if not ws.cell(row_idx, 1).value:
            continue
        row: dict[str, str] = {}
        for col_idx, key in enumerate(HEADERS, start=1):
            val = ws.cell(row_idx, col_idx).value
            row[key] = str(val).strip() if val is not None else ""
        # Normalize dropdown placeholders
        if row["for_each"].lower().startswith("(once"):
            row["for_each"] = ""
        if row["applies_when"].lower().startswith("(always"):
            row["applies_when"] = ""
        if not row["look_at"]:
            raise SystemExit(
                f"[excel] {row['id']}: Look at is blank. "
                f"Pick a cover field from the dropdown, or type the checkpoint name "
                f"EXACTLY as it appears on the sample report."
            )
        action = ACTION_MAP.get(row["rule_type"], row["rule_type"])
        where = f"{row['look_at']} {row['check_part']}".strip()
        blocks.append(
            {
                "id": row["id"],
                "row type": row["row_type"],
                "section": row["section"],
                "check": row["rule"],
                "where": where,
                "action": action,
                "param": row["value"],
                "when": row["applies_when"],
                "for each": row["for_each"],
                "example": row["example"],
            }
        )
    return blocks


def compile_excel_blocks(blocks: list[dict[str, str]]) -> tuple[list[dict], dict]:
    check_blocks: dict = {}
    macros: dict = {}
    rules: list = []

    for b in blocks:
        _validate_applies_when(b.get("id") or "?", b.get("when") or "")
        cb = block_to_check_block(b)
        rid = _slug(b.get("id") or "")
        check_blocks[rid] = cb
        if cb.get("row_type", "").lower().startswith("condition"):
            macros[rid] = cb  # full check_block — required for @C macro_ref
        else:
            rules.append(block_to_rule(b))

    for b in blocks:
        rid = _slug(b.get("id") or "")
        cb = check_blocks[rid]
        if cb.get("row_type", "").lower().startswith("condition"):
            continue
        when_raw = cb.get("when")
        if not isinstance(when_raw, str) or not when_raw.strip():
            cb["when"] = None
            continue
        if is_linked_when(when_raw):
            missing = [
                tok.strip().lstrip("@")
                for tok in re.findall(r"@[\w.]+", when_raw)
                if _slug(tok.strip().lstrip("@")) not in macros
            ]
            if missing:
                raise SystemExit(
                    f"[excel] {b.get('id')}: Applies when references missing Condition(s): "
                    f"{', '.join(missing)}. Add Row type=Condition (hidden) with that ID."
                )
            cb["when"] = _parse_linked_when(when_raw, macros)
        elif is_structured_when(when_raw):
            from harness_rules import when_to_predicate

            cb["when"] = when_to_predicate(when_raw, (b.get("where") or "").strip())
        else:
            # Should have been rejected in _validate_applies_when
            cb["when"] = "false"

    return rules, check_blocks


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path", nargs="?", default="data/clients/ribkoff/gi_authoring_ribkoff.xlsx")
    args = parser.parse_args()

    xlsx_path = Path(args.excel_path)
    if not xlsx_path.exists():
        raise SystemExit(f"No authoring Excel found at {xlsx_path}")

    blocks = load_excel_blocks(xlsx_path)
    rules, check_blocks = compile_excel_blocks(blocks)
    checkpoints = rules_to_checkpoints(rules, check_blocks=check_blocks)
    
    stem = xlsx_path.stem.replace("gi_authoring_", "")
    
    payload = {
        "meta": {"name": f"{stem} from Excel", "checkpoint_count": len(checkpoints), "source": str(xlsx_path)},
        "checkpoints": checkpoints,
    }

    out_path = Path(f"data/pipeline/checkpoints/{stem}_checkpoints.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    specs = resolve_specs(checkpoints)
    spec_path = Path(f"data/pipeline/checkspecs/{stem}_checkspecs.json")
    save_checkspecs(spec_path, specs)

    # Smoke: every checkpoint with a when must compile
    for cp in checkpoints:
        cb = cp.get("check_block") or {}
        if not cb:
            continue
        compile_block(cp, cb)

    n_cond = sum(1 for b in blocks if str(b.get("row type", "")).lower().startswith("condition"))
    print(
        f"Wrote {len(checkpoints)} checkpoints and {len(specs)} specs "
        f"from {xlsx_path.name} ({n_cond} Condition rows)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
