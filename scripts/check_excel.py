from openpyxl import load_workbook
wb = load_workbook("data/clients/ribkoff/gi_authoring_ribkoff.xlsx")
ws = wb["Checks"]
for i in range(4, 9):
    print([ws.cell(i, j).value for j in range(1, 12)])
