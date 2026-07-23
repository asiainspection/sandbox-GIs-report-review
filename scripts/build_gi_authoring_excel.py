#!/usr/bin/env python3
"""Build the ops GI authoring Excel (human-friendly dropdowns) + JSON schema.

Design goals (scale to ~200 clients, maintained by non-technical ops):
  - Split the old single "Where" string into two safe columns:
        "Look at"    = checkpoint / field name  (free text or suggestion)
        "Check part" = which part of it          (DROPDOWN, no typing suffixes)
    This kills the fragile "<name> <suffix>" string concatenation.
  - Add "For each" (iterator) so per-defect / per-SKU / per-PO rules are
    expressed structurally instead of being dumped on the LLM.
  - Rename technical terms to plain language:
        Action -> Rule type , Param -> Value , When -> Applies when
  - Keep one row = one atomic check.

Output:
  data/library/gi_authoring_template.xlsx
  data/library/gi_authoring_schema.json

Export path back to the compiler:
  where  = ("Look at" + " " + "Check part").strip()   # e.g. "Carton drop test photo count"
  action = Rule type -> harness action string          # see RULE_TYPES map below
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "data/library/gi_authoring_template.xlsx"
OUT_SCHEMA = ROOT / "data/library/gi_authoring_schema.json"

# ---------------------------------------------------------------------------
# RULE TYPES (operators) — human label -> (harness action, needs value?, kind, meaning)
# kind: "Deterministic" (Python decides, no AI) | "AI" | "Out-of-report"
# Keep the harness action strings in sync with harness_actions.md / harness_rules.py.
# ---------------------------------------------------------------------------
RULE_TYPES: list[dict] = [
    {"label": "Must be filled in", "action": "is present", "value": "no",
     "kind": "Deterministic", "meaning": "Field / checkpoint must not be blank."},
    {"label": "Must be in English", "action": "is in English", "value": "no",
     "kind": "Deterministic", "meaning": "No Chinese / CJK characters allowed."},
    {"label": "Must equal", "action": "equals", "value": "yes — the exact value",
     "kind": "Deterministic", "meaning": "Value must be exactly this (e.g. PENDING)."},
    {"label": "Must be one of", "action": "is one of", "value": "yes — comma list",
     "kind": "Deterministic", "meaning": "Value must be one of a fixed set (PASS, FAIL, N/A)."},
    {"label": "Must contain text", "action": "contains phrase", "value": "yes — exact phrase",
     "kind": "Deterministic", "meaning": "Mandatory wording must appear (e.g. 'Not applicable.')."},
    {"label": "Must not contain text", "action": "term must not appear", "value": "yes — banned term",
     "kind": "Deterministic", "meaning": "Banned wording must never appear (e.g. 'Golden Sample')."},
    {"label": "Number greater than", "action": "number greater than", "value": "yes — number (e.g. 0)",
     "kind": "Deterministic", "meaning": "Numeric field must be > value. Use with Defect count, quantities, photo count."},
    {"label": "Number less than", "action": "number less than", "value": "yes — number",
     "kind": "Deterministic", "meaning": "Numeric field must be < value."},
    {"label": "Defect type includes", "action": "defect type includes", "value": "yes — type words (e.g. dirt stain)",
     "kind": "Deterministic", "meaning": "True when report.defects contains a matching defect type. Use on Condition rows."},
    {"label": "At least N", "action": "at least N photos", "value": "yes — N",
     "kind": "Deterministic", "meaning": "Count (photos / items) must be >= N."},
    {"label": "At most N", "action": "at most N photos", "value": "yes — N",
     "kind": "Deterministic", "meaning": "Count must be <= N (0 = none allowed)."},
    {"label": "Ratio at least", "action": "ratio at least", "value": "yes — e.g. 0.8",
     "kind": "Deterministic", "meaning": "packed/ordered (etc.) ratio must be >= value."},
    {"label": "File name matches", "action": "filename matches", "value": "yes — pattern with *",
     "kind": "Deterministic", "meaning": "Attachment file name matches a simple pattern."},
    {"label": "AI document check", "action": "LLM quote then match", "value": "yes — what to look for",
     "kind": "AI", "meaning": "Uses LLM to find info in docs/comments. Flaky for long texts."},
    {"label": "AI yes/no (flag if false)", "action": "LLM yes/no", "value": "yes — one factual question",
     "kind": "AI", "meaning": "Ask the AI a yes/no question about the text. If NO, it fails."},
    {"label": "AI photo check (flag if false)", "action": "needs vision", "value": "yes — what the photo must show",
     "kind": "AI", "meaning": "Vision yes/no on photo content. Check part must be 'photo content'. If NO, it fails."},
    {"label": "Manual review", "action": "manual review", "value": "no",
     "kind": "Out-of-report", "meaning": "Evidence is outside the report (booking, SOP…). No auto-check."},
]
RULE_TYPE_LABELS = [r["label"] for r in RULE_TYPES]

# ---------------------------------------------------------------------------
# "Look at" — cover / external places (no Check part needed for these).
# Keep in sync with harness_where.md §2/§3 and harness_rules.COVER_WHERE.
# ---------------------------------------------------------------------------
COVER_PLACES = [
    "Factory address",
    "Factory name",
    "Supplier name",
    "Production site",
    "Inspector remark",
    "Overall result",
    "Inspection type",
    "PO reference",
    "Product name",
    "SKU",
    "Ordered quantity",
    "Packed quantity",
    "Produced quantity",
    "Unit",
    "Major AQL",
    "Minor AQL",
    "Critical AQL",
    "Workmanship result",
    "Defects",
    "Defect count",
    "Full report text",
    "All captions",
]

EXTERNAL_PLACES = [
    "Booking",
    "Spec sheet",
    "Email",
    "SOP",
]

# ---------------------------------------------------------------------------
# "Check part" — which part of a checklist checkpoint to read (DROPDOWN).
# Keep in sync with harness_where.md §1 and harness_rules.SUFFIX_TO_FIELD.
# (blank) is valid for cover / external places.
# ---------------------------------------------------------------------------
CHECK_PARTS = [
    "result",
    "comment",
    "values",
    "photo count",
    "photo caption",
    "photo content",
    "file name",
    "file content",
]

# ---------------------------------------------------------------------------
# "Repeat for" — iterator / scope. NOT product type, NOT PSI/DUPRO.
# Those belong in a Condition row (Look at = Inspection type / Product name).
# Blank = evaluate the rule once for the whole report.
# "each defect" is wired today. Others are structured roadmap.
# ---------------------------------------------------------------------------
FOR_EACH = [
    "",  # whole report (once)
    "each defect",
    "each SKU / reference",
    "each PO",
    "each photo",
    "each checkpoint in section",
]

# ---------------------------------------------------------------------------
# "Applies when" — blank OR link Condition IDs (@C1 AND @C2).
# Free prose is rejected at compile time. Conditions use the same Look at /
# Check part / Rule type / Value columns as Rules.
# ---------------------------------------------------------------------------
WHEN_PRESETS = [
    "",  # always
    "@C1",
    "@C1 AND @C2",
    "@C1 OR @C2",
    "NOT @C1",
]

# Soft suggestions for Value (Must equal / Must be one of / counts).
# Custom text still allowed (AI questions, filenames, comma lists).
VALUE_SUGGESTIONS = [
    # Results / verdicts
    "PASS",
    "FAIL",
    "Pending",
    "N/A",
    "NOT_APPLICABLE",
    "PASS, FAIL",
    "PASS, FAIL, Pending",
    # Inspection types (copy exact string from sample report when unsure)
    "Pre-Shipment Inspection",
    "During Production Inspection",
    # Common counts for At least N / At most N
    "0",
    "1",
    "2",
    "3",
    "5",
]

# ---------------------------------------------------------------------------
# Columns (final)
# ---------------------------------------------------------------------------
HEADERS = [
    "id",
    "row_type",
    "section",
    "rule",
    "applies_when",
    "for_each",
    "look_at",
    "check_part",
    "rule_type",
    "value",
    "example",
]
HEADER_LABELS = {
    "id": "ID",
    "row_type": "Row type",
    "section": "Section",
    "rule": "Rule — what must be true",
    "applies_when": "Applies when (blank = always, or @C1 AND @C2)",
    "for_each": "Repeat for (blank = once — NOT product/inspection type)",
    "look_at": "Look at — required: cover field OR exact checkpoint name from the report",
    "check_part": "Check part (blank only for cover fields like Factory address)",
    "rule_type": "Rule type",
    "value": "Value (blank if not needed)",
    "example": "Optional Example (Wrong: X · Right: X) — leave blank on Conditions",
}

# Example shows the production pattern: Condition rows + @refs on Rules.
EXAMPLE_ROWS = [
    {
        "id": "C1", "row_type": "Condition (hidden)", "section": "Conditions",
        "rule": "Any defects found",
        "applies_when": "", "for_each": "",
        "look_at": "Defect count", "check_part": "",
        "rule_type": "Number greater than", "value": "0",
        "example": "",
    },
    {
        "id": "A.2.1", "row_type": "Rule", "section": "QIMA2 — Inspector remark / defects by PO",
        "rule": "Inspector remark should list all defects found broken down by PO number with a total defect %",
        "applies_when": "@C1", "for_each": "",
        "look_at": "Inspector remark", "check_part": "",
        "rule_type": "AI yes/no (flag if false)",
        "value": "Does this remark list defects broken down by PO number and include a total defect percentage?",
        "example": "Wrong: \"3 defects found\" · Right: \"PO 081901: 3 defects … Total: 6.00%\"",
    },
    {
        "id": "C2", "row_type": "Condition (hidden)", "section": "Conditions",
        "rule": "Carton drop test passed",
        "applies_when": "", "for_each": "",
        "look_at": "Carton drop test", "check_part": "result",
        "rule_type": "Must equal", "value": "PASS",
        "example": "",
    },
    {
        "id": "A.5.1", "row_type": "Rule", "section": "QIMA5 — Carton drop test photos",
        "rule": "Carton-drop photos should not be included when the test passed",
        "applies_when": "@C2", "for_each": "",
        "look_at": "Carton drop test", "check_part": "photo count",
        "rule_type": "At most N", "value": "0",
        "example": "Wrong: drop photos on Pass · Right: no photos on Pass",
    },
    {
        "id": "C3", "row_type": "Condition (hidden)", "section": "Conditions",
        "rule": "Inspection is PSI",
        "applies_when": "", "for_each": "",
        "look_at": "Inspection type", "check_part": "",
        "rule_type": "Must equal", "value": "Pre-Shipment Inspection",
        "example": "",
    },
    {
        "id": "C4", "row_type": "Condition (hidden)", "section": "Conditions",
        "rule": "Overall result is FAIL",
        "applies_when": "", "for_each": "",
        "look_at": "Overall result", "check_part": "",
        "rule_type": "Must equal", "value": "FAIL",
        "example": "",
    },
    {
        "id": "X.1", "row_type": "Rule", "section": "Example — two conditions",
        "rule": "Example rule that only runs on PSI reports that FAILED",
        "applies_when": "@C3 AND @C4", "for_each": "",
        "look_at": "Inspector remark", "check_part": "",
        "rule_type": "Must be filled in", "value": "",
        "example": "Wrong: blank remark · Right: filled remark",
    },
]

BLANK_ROWS = 40


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _thin_border(cell) -> None:
    side = Side(style="thin", color="B0B0B0")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _build_lists(wb: Workbook):
    lists = wb.active
    lists.title = "Lists"
    cols = {
        "A": ("rule_type", RULE_TYPE_LABELS),
        "B": ("cover_place", COVER_PLACES),
        "C": ("external_place", EXTERNAL_PLACES),
        "D": ("check_part", CHECK_PARTS),
        "E": ("repeat_for", [f or "(once — blank)" for f in FOR_EACH]),
        "F": ("applies_when", [w or "(always — blank)" for w in WHEN_PRESETS]),
        "G": ("look_at_suggestions", COVER_PLACES + EXTERNAL_PLACES),
        "H": ("value_suggestions", VALUE_SUGGESTIONS),
    }
    for letter, (header, values) in cols.items():
        cell = lists[f"{letter}1"]
        cell.value = header
        _style_header(cell)
        for i, v in enumerate(values, start=2):
            lists[f"{letter}{i}"] = v
        lists.column_dimensions[letter].width = 30
    return lists


def _build_checks(wb: Workbook, rows_data: list[dict[str, str]]):
    checks = wb.create_sheet("Checks", 0)
    n_cols = len(HEADERS)
    last_col = get_column_letter(n_cols)

    checks["A1"] = "GI checks — one row = one atomic check"
    checks["A1"].font = Font(bold=True, size=14, color="1F4E79")
    checks.merge_cells(f"A1:{last_col}1")
    checks["A2"] = (
        "ONE ROW = ONE CHECK. Rule = what we check. Condition = when it applies (hidden; link with @C1). "
        "Applies when: blank OR @C1 / @C1 AND @C2 (dropdown) — no free sentences. "
        "Look at: always fill — pick a cover field from the list, OR type the checkpoint name EXACTLY "
        "as it appears on the sample report (e.g. Carton drop test). "
        "Value: for Must equal / Must be one of, prefer the Value dropdown (PASS, Pre-Shipment Inspection…). "
        "Custom Values still OK (AI questions, file patterns). "
        "Repeat for = each defect/SKU/PO only — not PSI or product type (use a Condition for those)."
    )
    checks["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    checks.merge_cells(f"A2:{last_col}2")
    checks.row_dimensions[2].height = 44

    for col, key in enumerate(HEADERS, start=1):
        cell = checks.cell(3, col, HEADER_LABELS[key])
        _style_header(cell)
        _thin_border(cell)

    all_rows = rows_data + [{h: "" for h in HEADERS} for _ in range(BLANK_ROWS)]
    example_fill = PatternFill("solid", fgColor="E8F5E9")
    for r_i, row in enumerate(all_rows, start=4):
        for c_i, key in enumerate(HEADERS, start=1):
            cell = checks.cell(r_i, c_i, row.get(key, ""))
            _thin_border(cell)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_i < 4 + len(rows_data):
                cell.fill = example_fill

    last_data_row = 3 + len(all_rows)
    widths = {
        "id": 7, "row_type": 18, "section": 13, "rule": 46, "applies_when": 28, "for_each": 22,
        "look_at": 42, "check_part": 22, "rule_type": 20, "value": 22, "example": 34,
    }
    for idx, key in enumerate(HEADERS, start=1):
        checks.column_dimensions[get_column_letter(idx)].width = widths[key]
    checks.row_dimensions[3].height = 34
    checks.freeze_panes = "A4"

    def col_letter(key: str) -> str:
        return get_column_letter(HEADERS.index(key) + 1)

    # Strict dropdown: Row type
    dv_row = DataValidation(
        type="list", formula1='"Rule,Condition (hidden)"',
        allow_blank=False, showDropDown=False, showErrorMessage=True,
    )
    dv_row.error = "Pick Rule or Condition."
    dv_row.errorTitle = "Invalid Row type"
    checks.add_data_validation(dv_row)
    dv_row.add(f"{col_letter('row_type')}4:{col_letter('row_type')}{last_data_row}")

    # Strict dropdown: Rule type
    dv_rule = DataValidation(
        type="list", formula1=f"Lists!$A$2:$A${len(RULE_TYPE_LABELS) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=True,
    )
    dv_rule.error = "Pick a Rule type from the list."
    dv_rule.errorTitle = "Invalid Rule type"
    dv_rule.promptTitle = "Rule type"
    dv_rule.prompt = "Pick one. Deterministic first; AI only when needed."
    checks.add_data_validation(dv_rule)
    dv_rule.add(f"{col_letter('rule_type')}4:{col_letter('rule_type')}{last_data_row}")

    # Look at: suggestions from cover fields; custom text allowed for exact checkpoint names.
    # Blank is allowed in Excel (so typing a name is not blocked) but compile requires a value.
    dv_look = DataValidation(
        type="list", formula1=f"Lists!$G$2:$G${len(COVER_PLACES) + len(EXTERNAL_PLACES) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=False,
    )
    dv_look.promptTitle = "Look at (required)"
    dv_look.prompt = (
        "Always fill this. Option A: pick a cover field (Factory address, Inspection type…). "
        "Option B: type the checkpoint name EXACTLY as on the sample report "
        "(e.g. Carton drop test) — spelling must match. Then pick Check part."
    )
    checks.add_data_validation(dv_look)
    dv_look.add(f"{col_letter('look_at')}4:{col_letter('look_at')}{last_data_row}")

    # Applies when: blank or @Condition refs only
    dv_when = DataValidation(
        type="list", formula1=f"Lists!$F$2:$F${len(WHEN_PRESETS) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=True,
    )
    dv_when.error = (
        "Pick blank (always) or an @C link from the list. "
        "Do not type free text — add a Condition row, then link it here."
    )
    dv_when.errorTitle = "Applies when"
    dv_when.promptTitle = "Applies when"
    dv_when.prompt = (
        "Blank = always. Else @C1 or @C1 AND @C2. "
        "Define C1/C2 as Row type = Condition (hidden)."
    )
    checks.add_data_validation(dv_when)
    dv_when.add(f"{col_letter('applies_when')}4:{col_letter('applies_when')}{last_data_row}")

    # Strict dropdown: Check part
    dv_part = DataValidation(
        type="list", formula1=f"Lists!$D$2:$D${len(CHECK_PARTS) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=True,
    )
    dv_part.error = "Pick a Check part, or leave blank for a cover field."
    dv_part.errorTitle = "Invalid Check part"
    dv_part.promptTitle = "Check part"
    dv_part.prompt = "Which part of the checkpoint? Blank for a cover field."
    checks.add_data_validation(dv_part)
    dv_part.add(f"{col_letter('check_part')}4:{col_letter('check_part')}{last_data_row}")

    # Strict dropdown: Repeat for (iterator — not product/inspection type)
    dv_each = DataValidation(
        type="list", formula1=f"Lists!$E$2:$E${len(FOR_EACH) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=True,
    )
    dv_each.error = "Pick a Repeat-for option, or leave blank to run once."
    dv_each.errorTitle = "Repeat for"
    dv_each.promptTitle = "Repeat for"
    dv_each.prompt = (
        "Blank = once for the whole report. each defect = same check for every defect. "
        "For PSI or product type, use a Condition row instead."
    )
    checks.add_data_validation(dv_each)
    dv_each.add(f"{col_letter('for_each')}4:{col_letter('for_each')}{last_data_row}")

    # Soft Value suggestions (Must equal / Must be one of / counts). Custom text allowed.
    dv_val = DataValidation(
        type="list", formula1=f"Lists!$H$2:$H${len(VALUE_SUGGESTIONS) + 1}",
        allow_blank=True, showDropDown=False, showErrorMessage=False,
    )
    dv_val.promptTitle = "Value (suggestions)"
    dv_val.prompt = (
        "For Must equal / Must be one of: pick PASS, FAIL, Pre-Shipment Inspection, etc. "
        "For At least/At most N: pick 0,1,2,3,5. "
        "You may type custom values (AI questions, file patterns, comma lists)."
    )
    checks.add_data_validation(dv_val)
    dv_val.add(f"{col_letter('value')}4:{col_letter('value')}{last_data_row}")

    return checks


def _build_operators(wb: Workbook):
    ops = wb.create_sheet("Operators")
    headers = ["Rule type", "What it means", "Needs a value?", "Kind", "Maps to (harness action)"]
    for c, h in enumerate(headers, start=1):
        cell = ops.cell(1, c, h)
        _style_header(cell)
        _thin_border(cell)
    kind_fill = {
        "Deterministic": PatternFill("solid", fgColor="E8F5E9"),
        "AI": PatternFill("solid", fgColor="FFF3E0"),
        "Out-of-report": PatternFill("solid", fgColor="ECEFF1"),
    }
    for r, rt in enumerate(RULE_TYPES, start=2):
        row = [rt["label"], rt["meaning"], rt["value"], rt["kind"], rt["action"]]
        for c, v in enumerate(row, start=1):
            cell = ops.cell(r, c, v)
            _thin_border(cell)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = kind_fill[rt["kind"]]
    widths = [20, 52, 24, 16, 26]
    for i, w in enumerate(widths, start=1):
        ops.column_dimensions[get_column_letter(i)].width = w
    ops.freeze_panes = "A2"
    return ops


def _build_readme(wb: Workbook):
    readme = wb.create_sheet("Readme")
    readme["A1"] = "How to author a GI check (ops — production)"
    readme["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "IN ONE SENTENCE",
        "---------------",
        "A Rule is what must be true. A Condition is when that Rule runs. Link them with @C1.",
        "",
        "LOOK AT (always fill — do not leave blank on a real row)",
        "-------------------------------------------------------",
        "  Cover field:   pick from the dropdown (Factory address, Inspection type, Overall result…).",
        "                 Leave Check part blank.",
        "  Checkpoint:    open the sample report and COPY the checklist title exactly",
        "                 (e.g. Carton drop test — not 'carton drop' or 'drop test').",
        "                 Then pick Check part: result / photo count / comment / …",
        "  Wrong: blank Look at · Wrong: approximate name · Right: exact report wording.",
        "",
        "APPLIES WHEN",
        "------------",
        "  Blank = this Rule always runs.",
        "  Or pick @C1 / @C1 AND @C2 from the dropdown after you create Condition rows.",
        "  Do not write sentences like 'when the drop test passed' — that is rejected.",
        "",
        "HOW TO MAKE A CONDITION (e.g. only for PSI)",
        "-------------------------------------------",
        "  1. New row, Row type = Condition (hidden), ID = C1",
        "  2. Look at = Inspection type, Rule type = Must equal, Value = Pre-Shipment Inspection",
        "  3. Leave Optional Example blank on Condition rows",
        "  4. On every PSI-only Rule: Applies when = @C1",
        "",
        "  Same idea for product: Look at = Product name, Must equal / Must contain text.",
        "",
        "VALUE (Must equal / Must be one of)",
        "----------------------------------",
        "  Prefer the Value dropdown: PASS, FAIL, Pending, N/A, NOT_APPLICABLE,",
        "  Pre-Shipment Inspection, During Production Inspection, or PASS, FAIL, Pending.",
        "  Copy Inspection type EXACTLY from the sample report (not the nickname 'PSI').",
        "  Custom Values are OK for AI questions and file-name patterns.",
        "",
        "REPEAT FOR",
        "----------",
        "  Blank = once. each defect = same check for every defect.",
        "  Not for PSI or product type — that is a Condition.",
        "",
        "TWO CONDITIONS TOGETHER (PSI and FAIL)",
        "--------------------------------------",
        "  C3 Condition  Inspection type = Pre-Shipment Inspection",
        "  C4 Condition  Overall result = FAIL",
        "  Rule          Applies when = @C3 AND @C4",
        "",
        "Prefer Must equal / At least N before AI yes/no.",
    ]
    for i, line in enumerate(lines, start=2):
        readme[f"A{i}"] = line
    readme.column_dimensions["A"].width = 104
    return readme


def build_workbook(rows_data: list[dict[str, str]] = None) -> Workbook:
    if rows_data is None:
        rows_data = EXAMPLE_ROWS
    wb = Workbook()
    _build_lists(wb)          # sheet index 0 initially (active) -> renamed Lists
    _build_checks(wb, rows_data)         # inserted at index 0 (first tab)
    _build_operators(wb)
    _build_readme(wb)
    # Order tabs: Checks, Operators, Readme, Lists
    wb.move_sheet("Lists", offset=len(wb.sheetnames))
    return wb


def build_schema() -> dict:
    return {
        "schema_version": 2,
        "title": "GI check authoring",
        "description": (
            "One row = one atomic check. 'Look at' + 'Check part' replace the old single "
            "'where' string (export joins them). 'For each' is the iterator/scope. "
            "'Rule type' is a closed enum mapping to harness actions."
        ),
        "source_docs": [
            "data/library/harness_actions.md",
            "data/library/harness_where.md",
        ],
        "fields": [
            {"key": "id", "label": "ID", "type": "string", "required": True},
            {"key": "section", "label": "Section", "type": "string", "required": False},
            {"key": "rule", "label": "Rule — what must be true", "type": "string",
             "required": True, "ui": "textarea"},
            {"key": "applies_when", "label": "Applies when", "type": "string",
             "required": False, "ui": "select",
             "enum": [w for w in WHEN_PRESETS if w],
             "help": "Blank = always. Else @C1 or @C1 AND @C2 linking Condition rows. No free prose."},
            {"key": "for_each", "label": "Repeat for", "type": "enum",
             "required": False, "ui": "select", "enum": [f for f in FOR_EACH if f],
             "help": "Iterator only (each defect/SKU/PO). NOT product type or PSI — those are Condition rows."},
            {"key": "look_at", "label": "Look at", "type": "string",
             "required": True, "ui": "combobox",
             "suggestions": COVER_PLACES + EXTERNAL_PLACES,
             "pattern_hint": "Required. Cover field from list, or checkpoint name copied exactly from the sample report."},
            {"key": "check_part", "label": "Check part", "type": "enum",
             "required": False, "ui": "select", "enum": CHECK_PARTS,
             "help": "Which part of a checklist checkpoint. Blank for a cover field."},
            {"key": "rule_type", "label": "Rule type", "type": "enum",
             "required": True, "ui": "select", "enum": RULE_TYPE_LABELS,
             "actions": {r["label"]: r["action"] for r in RULE_TYPES}},
            {"key": "value", "label": "Value", "type": "string",
             "required": False, "ui": "combobox",
             "suggestions": VALUE_SUGGESTIONS,
             "help": "For Must equal / Must be one of prefer PASS, FAIL, Pre-Shipment Inspection…. Soft list; custom OK for AI questions."},
            {"key": "example", "label": "Example", "type": "string",
             "required": False, "ui": "textarea"},
        ],
        "export": {
            "csv_columns": HEADERS,
            "sheet": "Checks",
            "where_from": ["look_at", "check_part"],
            "action_from": "rule_type",
        },
    }


def main() -> int:
    wb = build_workbook()
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    OUT_SCHEMA.write_text(json.dumps(build_schema(), indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_SCHEMA}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
