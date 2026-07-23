"""Build report-team findings Excel from PolicyGuard review flags.

Columns the team reads (keep short):
  What was checked  → always "GI Instructions"
  Checkpoint        → where locus (e.g. "Carton Drop Test")
  Original Content  → exact report value extracted
  Non-confirmities  → Rules: … Inconsistency: …
  Suggested actions → short verb-first fix
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FINDINGS_HEADERS = [
    "Order #",
    "Empty",
    "1",
    "What was checked",
    "Checkpoint",
    "Original Content",
    "Non-confirmities",
    "Suggested actions",
    "Finding Verdict",
    "Remark",
]

WHAT_WAS_CHECKED = "GI Instructions"
WHAT_WAS_CHECKED_UNABLE = "GI Instructions - Unable to check"
# Original Content fallbacks — plain language for the report team (not engineer jargon).
ORIGINAL_MISSING = "This value is not in the report (empty or not filled in)."
ORIGINAL_NOT_FOUND = "Could not find this field in the report."
ORIGINAL_SUPPRESSED = "Report value present — see Checkpoint column (technical quote hidden)."
MANUAL_ATTACHMENT_ACTION = (
    "Please manually check, attachments processing is not supported for the moment."
)
MANUAL_PHOTO_ACTION = (
    "Please manually check, image content processing is not supported for the moment."
)

_SELECTOR_RE = re.compile(
    r"checklist\.(?P<node>[a-z0-9_]+)\.(?P<field>photo_count|caption_count|comment|result|attachment_filenames|photo_content|attachment_content)"
    r"(?:=(?P<val>[^,;]+))?",
    re.I,
)
_REPORT_SEL_RE = re.compile(
    r"report\.(?P<field>[a-z0-9_]+)(?:=(?P<val>.*))?$",
    re.I | re.S,
)
_MAX_RE = re.compile(r"\bmax=(\d+)", re.I)
_MIN_RE = re.compile(r"\bmin=(\d+)", re.I)
_STYLE_RE = re.compile(r"missing style\s+(\S+)", re.I)
_DASH_SPLIT = re.compile(r"\s+[—–]\s+")


def order_id_for_report(report_path: Path, report: dict[str, Any] | None = None) -> str:
    """Prefer Q-number from filename; fall back to inspection id."""
    stem = report_path.stem.replace("_flawed", "").replace("_llm", "")
    q = re.match(r"(Q\d{9,10})", stem, re.I)
    if q:
        return q.group(1)
    if report is None and report_path.exists():
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
        except Exception:
            report = None
    if isinstance(report, dict):
        for key in ("inspectionId", "reportId", "qspInspectionId"):
            val = report.get(key)
            if val:
                return str(val)
    return stem


def requirements_by_id(checkpoints: list[dict[str, Any]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for cp in checkpoints:
        cid = str(cp.get("id") or "")
        req = str(cp.get("requirement") or cp.get("what_to_check") or "").strip()
        if cid:
            out[cid] = req
    return out


def checkpoint_meta_by_id(checkpoints: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """id → requirement / fail_if / where / field_location for handoff wording."""
    out: dict[str, dict[str, Any]] = {}
    for cp in checkpoints:
        cid = str(cp.get("id") or "")
        if not cid:
            continue
        fail_if = cp.get("fail_if") or []
        if isinstance(fail_if, str):
            fail_if = [fail_if]
        requirement = str(cp.get("requirement") or cp.get("what_to_check") or "").strip()
        where = (cp.get("check_block") or {}).get("where")
        out[cid] = {
            "requirement": requirement,
            "fail_if": [str(x).strip() for x in fail_if if str(x).strip()],
            "section": str(cp.get("section") or "").strip(),
            "severity": str(cp.get("severity") or "").strip(),
            "where": where,
            "field_location": _field_location_from_requirement(requirement, where),
            "rule_short": _rule_short_from_requirement(requirement, fail_if),
        }
    return out


def _node_label(node: str) -> str:
    return node.replace("_", " ").strip().title()


def _clip(text: str, n: int = 120) -> str:
    text = re.sub(r"\s+", " ", (text or "").strip())
    if len(text) <= n:
        return text
    return text[: n - 1].rstrip() + "…"


def _field_location_from_requirement(requirement: str, where: Any = None) -> str:
    """Human locus for Checkpoint column — prefer Field/Location, else where binding."""
    req = (requirement or "").strip()
    if req:
        parts = _DASH_SPLIT.split(req, maxsplit=1)
        head = (parts[0] or "").strip()
        # Drop trailing "checklist / report photos" noise when it's the whole head.
        if head and len(head) <= 70:
            return head
    label = _label_from_where(where)
    return label or "Report"


def _rule_short_from_requirement(requirement: str, fail_if: list[Any] | None = None) -> str:
    """Short GI rule for Non-conformities (What-to-check / fail_if), not the locus."""
    for item in fail_if or []:
        text = str(item)
        if "Report does not satisfy:" in text:
            return _clip(text.split("Report does not satisfy:", 1)[-1].strip(), 110)
    req = (requirement or "").strip()
    if not req:
        return ""
    parts = _DASH_SPLIT.split(req, maxsplit=2)
    # requirement often: Field — section — What to check
    body = parts[-1].strip() if parts else req
    return _clip(body, 110)


def _label_from_where(where: Any) -> str | None:
    if not where:
        return None
    entry = where[0] if isinstance(where, list) and where else where
    if isinstance(entry, str):
        text = entry.strip()
        if text.startswith("checklist."):
            parts = text.split(".")
            if len(parts) >= 2:
                return _node_label(parts[1])
        if text.startswith("report."):
            return _node_label(text.split(".", 1)[1])
        if text.startswith("workmanship."):
            return _node_label(text.split(".", 1)[1])
        if text.startswith("product."):
            return "Product quantity"
        if text.startswith("out_of_report"):
            return "Out of report"
        return _clip(text, 50)
    if isinstance(entry, dict):
        kind = str(entry.get("kind") or entry.get("type") or "").lower()
        match = entry.get("match") or []
        if isinstance(match, str):
            match = [match]
        if match:
            return _clip(" ".join(str(m) for m in match).replace("_", " ").title(), 50)
        field = str(entry.get("field") or "").strip()
        if kind and field:
            return f"{kind.title()} · {field}"
    return None


def checkpoint_column(flag: dict[str, Any], meta: dict[str, Any] | None) -> str:
    """Where locus shown to the team (not the checkpoint id)."""
    if meta and meta.get("field_location"):
        return str(meta["field_location"])
    ev = str(flag.get("evidence") or "")
    m = _SELECTOR_RE.search(ev)
    if m:
        return _node_label(m.group("node"))
    rm = _REPORT_SEL_RE.search(ev.split(";")[0].strip())
    if rm:
        return _node_label(rm.group("field"))
    where = (meta or {}).get("where")
    return _label_from_where(where) or _clip(str(flag.get("section") or "Report"), 50)


def _quote(text: str, n: int = 100) -> str:
    body = _clip(text, n)
    if not body:
        return "(blank)"
    return f'"{body}"'


def original_content(flag: dict[str, Any], report_snippet: str | None = None) -> str:
    """Exact report content the rule looked at — short.

    Never surface LLM Requirement echoes or raw atom quotes as the report value.
    Prefer a live report snippet; else deterministic selector evidence only.
    Plain atom evidence (address / remark quote) is shown when it is substantive.
    """
    if report_snippet and report_snippet.strip():
        clipped = _clip(report_snippet.strip(), 140)
        if not _looks_like_llm_junk(clipped):
            return clipped

    ev = str(flag.get("evidence") or "").strip()
    if not ev:
        return ORIGINAL_NOT_FOUND
    if _looks_like_status_phrase(ev):
        return _humanize_status_evidence(ev)
    if _looks_like_llm_junk(ev):
        return ORIGINAL_SUPPRESSED

    if "report.global_remark" in ev:
        m = re.search(r"report\.global_remark=(.+)$", ev, re.S)
        if m:
            body = m.group(1).strip()
            if body.lower() in ("true", "false", ""):
                return "Inspector remark: (blank)" if "blank=True" in ev or not body else body
            return f"Inspector remark: {_quote(body, 100)}"
        if "blank=True" in ev:
            return "Inspector remark: (blank)"

    style = _STYLE_RE.search(ev)
    if style:
        return f"Attachments: style {style.group(1)} missing"

    m = _SELECTOR_RE.search(ev)
    if m:
        field = m.group("field")
        val = (m.group("val") or "").strip()
        node = _node_label(m.group("node"))
        if field == "photo_count":
            return f"{node}: {val or '?'} photo(s)"
        if field == "caption_count":
            return f"{node}: {val or '?'} caption(s)"
        if field == "comment":
            if "blank=True" in ev or not val:
                return f"{node} comment: (blank)"
            return f"{node} comment: {_quote(val, 90)}"
        if field == "result":
            return f"{node} result: {val or '(empty)'}"
        if field == "attachment_filenames":
            return f"{node} files: {_clip(val or '(none)', 90)}"

    rm = _REPORT_SEL_RE.search(ev.split(";")[0].strip())
    if rm:
        field = _node_label(rm.group("field"))
        val = (rm.group("val") or "").strip()
        if "blank=True" in ev or val == "":
            return f"{field}: (blank)"
        if val.lower() in ("true", "false"):
            return f"{field}: {val}"
        return f"{field}: {_quote(val, 90)}"

    # Deterministic defect-filter / count evidence is OK to show
    if ev.startswith("defects_name_any") or "photo_count=" in ev or "max=" in ev or "min=" in ev:
        return _clip(ev, 140)

    # Atom / LLM evidence is often a raw quote without selector=value.
    # If it looks like real report text, surface it (better than a missing-value fallback).
    if _looks_like_report_quote(ev):
        return _clip(ev, 140)

    return ORIGINAL_MISSING


def _humanize_status_evidence(text: str) -> str:
    """Map engine status phrases → report-team wording."""
    low = (text or "").strip().lower()
    if "required field blank" in low or "bound field blank" in low:
        return ORIGINAL_MISSING
    if "bound field missing" in low or "binding unresolved" in low:
        return ORIGINAL_NOT_FOUND
    if "no filenames" in low:
        return "No attachment file name found in the report."
    if "vision leaf not evaluated" in low:
        return "Photo content was not checked (vision off or no photos)."
    if "insufficient evidence" in low:
        return ORIGINAL_MISSING
    return ORIGINAL_MISSING


def _looks_like_status_phrase(text: str) -> bool:
    low = (text or "").strip().lower()
    if not low:
        return False
    phrases = (
        "required field blank",
        "bound field missing",
        "bound field blank",
        "excuse field blank",
        "vision leaf not evaluated",
        "insufficient evidence",
        "no filenames",
        "binding unresolved",
    )
    return any(low == p or low.startswith(p) for p in phrases)


def _looks_like_report_quote(text: str) -> bool:
    """True when evidence is likely a copied report string (atom path)."""
    t = (text or "").strip()
    if len(t) < 8:
        return False
    if _looks_like_status_phrase(t) or _looks_like_llm_junk(t):
        return False
    if "=" in t and any(
        t.lower().startswith(p)
        for p in ("checklist.", "report.", "product.", "workmanship.", "atom.", "vision.")
    ):
        return False
    return True


def _looks_like_llm_junk(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return False
    low = t.lower()
    if low.startswith("requirement:"):
        return True
    if "requirement:" in low and "defects —" in low:
        return True
    if t.startswith("{") and "'name':" in t:
        return True  # raw defect dict dump
    if t.startswith("{") and '"name":' in t:
        return True
    if low.startswith("answer only from the bound field"):
        return True
    return False


def non_conformity(
    flag: dict[str, Any],
    meta: dict[str, Any] | None,
    report_snippet: str | None = None,
) -> str:
    """Rules: <GI>. Inconsistency: <expected vs report>."""
    rule = _clip(str((meta or {}).get("rule_short") or (meta or {}).get("requirement") or ""), 100)
    if not rule:
        rule = "See GI for this checkpoint"

    # Use the same enriched snippet as Original Content so columns stay consistent.
    report_has = original_content(flag, report_snippet)
    expected = _expected_hint(flag, meta)
    if expected:
        inconsistency = f"expected {expected}; report has {report_has}"
    else:
        inconsistency = f"report has {report_has}"
    return f"Rules: {rule}. Inconsistency: {inconsistency}."


def _expected_hint(flag: dict[str, Any], meta: dict[str, Any] | None) -> str:
    ev = str(flag.get("evidence") or "")
    max_m = _MAX_RE.search(ev)
    min_m = _MIN_RE.search(ev)
    m = _SELECTOR_RE.search(ev)
    if m and m.group("field") == "photo_count":
        if max_m and max_m.group(1) == "0":
            return "0 photos when PASS"
        if max_m:
            return f"at most {max_m.group(1)} photo(s)"
        if min_m:
            return f"at least {min_m.group(1)} photo(s)"
    if m and m.group("field") == "comment" and ("blank=True" in ev or not (m.group("val") or "").strip()):
        return "required comment text"
    if "global_remark" in ev and ("blank=True" in ev or re.search(r"global_remark=\s*$", ev)):
        return "defect list by PO + %"
    if _STYLE_RE.search(ev):
        return "Measurement Chart-<style>.xlsx"
    # Prefer first fail_if short clause when present
    for item in (meta or {}).get("fail_if") or []:
        text = str(item).strip()
        if text and "Report does not satisfy" not in text and len(text) < 80:
            return _clip(text, 70)
    return ""


def suggested_actions(flag: dict[str, Any], meta: dict[str, Any] | None) -> str:
    """Short verb-first action the inspector / report editor can do."""
    ev = str(flag.get("evidence") or "")
    cid = str(flag.get("checkpoint_id") or "")
    locus = checkpoint_column(flag, meta)
    node_m = _SELECTOR_RE.search(ev)
    field = node_m.group("field") if node_m else ""
    max_m = _MAX_RE.search(ev)
    max_n = int(max_m.group(1)) if max_m else None
    min_m = _MIN_RE.search(ev)
    min_n = int(min_m.group(1)) if min_m else None

    if "global_remark" in ev or cid in ("1_2_1", "1_3_2", "a_2_1"):
        return 'Add Inspector Remark listing defects by PO (format: "PO xxx: X pcs / Total: X%").'

    style = _STYLE_RE.search(ev)
    if style or cid in ("5_1_1", "a_16_1"):
        style_n = style.group(1) if style else "<style>"
        return f'Replace measurement file with "Measurement Chart-{style_n}.xlsx".'

    if field == "photo_count" and max_n == 0:
        return f"Remove photos from {locus} (allowed only when the test FAILS)."
    if field == "photo_count" and max_n is not None:
        return f"Keep at most {max_n} photo(s) on {locus}; remove extras."
    if field == "photo_count" and min_n is not None:
        return f"Upload at least {min_n} photo(s) on {locus}."
    if field == "caption_count" or cid == "a_9_2" or "required field blank" in ev.lower():
        if "caption" in locus.lower() or cid == "a_9_2":
            return f"Add a caption on every photo under {locus}."
        return f"Fill the missing value on {locus}."
    if field == "comment" and ("blank=True" in ev or re.search(r"comment=\s*(;|$)", ev)):
        return f"Add the required comment on {locus}."
    if field == "result" or cid == "a_6_1":
        return f"Correct the result on {locus} per the GI."
    if "factory_address" in ev or "chinese" in ev.lower() or cid in ("a_1_1", "a_1_2"):
        if cid == "a_1_2":
            return "Rewrite factory address in English (street, city, country, postal code)."
        return "Add the complete factory address including postal code."
    if "missing" in ev.lower():
        return f"Upload the missing attachment/photo for {locus}."
    if "blank=True" in ev:
        return f"Fill the missing value on {locus}."

    return f"Correct {locus} to match the GI rule."


def enrich_report_snippet(
    report: dict[str, Any] | None,
    evidence: str,
    *,
    meta: dict[str, Any] | None = None,
    spec: dict[str, Any] | None = None,
) -> str | None:
    """Pull a live checklist/report snippet from the report JSON when possible."""
    if not report:
        return None
    ev = evidence or ""
    m = _SELECTOR_RE.search(ev)
    if not m:
        if "defects_name_any" in ev or ev.lower().startswith("defects:"):
            return _defect_rows_snippet(report)
        if "all_captions" in ev or "all photo captions" in ev.lower():
            return "All photo captions"
        # Atom evidence with no selector — resolve via authored where bindings FIRST
        # (before loose keyword matches that false-positive on remark text).
        from_where = _snippet_from_where_bindings(report, meta=meta, spec=spec)
        if from_where:
            return from_where
        if "global_remark" in ev:
            remark = (report.get("result") or {}).get("globalRemark") or {}
            msg = str(remark.get("message") or "").strip()
            return f"Inspector remark: {_quote(msg, 100)}" if msg else "Inspector remark: (blank)"
        if "factory_address" in ev:
            for key in ("factoryAddress", "factory_address", "address"):
                val = report.get(key)
                if val:
                    return f"Factory address: {_quote(str(val), 100)}"
        return None

    target = normalize_node(m.group("node"))
    field = m.group("field")
    for el in _iter_checklist_elements(report):
        name = _element_name(el)
        if re.sub(r"[^a-z0-9]", "", name.lower()) != re.sub(r"[^a-z0-9]", "", target):
            if normalize_node(name) != target:
                continue
        result = el.get("result")
        imgs = len(el.get("images") or [])
        comment = ""
        c = el.get("comment") or {}
        if isinstance(c, dict):
            comment = str(c.get("message") or "").strip()
        if field == "comment":
            return f"{name} comment: {_quote(comment, 90) if comment else '(blank)'} (result={result})"
        if field in ("photo_count", "caption_count"):
            cap = sum(1 for im in (el.get("images") or []) if (im.get("caption") or "").strip())
            return f"{name}: photos={imgs}, captions={cap}, result={result}"
        if field == "result":
            return f"{name} result={result}"
        if field == "photo_content":
            caps = [
                str(im.get("caption") or "").strip()
                for im in (el.get("images") or [])
                if str(im.get("caption") or "").strip()
            ]
            return f"{name}: photos={imgs}, captions={_clip(', '.join(caps) if caps else '(none)', 80)}"
        if field == "attachment_filenames":
            files = el.get("attachments") or el.get("files") or []
            names = []
            for f in files:
                if isinstance(f, dict):
                    names.append(str(f.get("name") or f.get("fileName") or f.get("filename") or ""))
                else:
                    names.append(str(f))
            names = [n for n in names if n]
            return f"{name} files: {_clip(', '.join(names) if names else '(none)', 90)}"
        return f"{name}: result={result}, photos={imgs}"
    return None


def _snippet_from_where_bindings(
    report: dict[str, Any],
    *,
    meta: dict[str, Any] | None = None,
    spec: dict[str, Any] | None = None,
) -> str | None:
    """Resolve authored where → live semantic values for Original Content."""
    where = None
    if meta and meta.get("where"):
        where = meta.get("where")
    elif spec and (spec.get("where_bindings") or spec.get("where")):
        where = spec.get("where_bindings") or spec.get("where")
    if not where:
        return None
    try:
        from fact_index import build_fact_index
        from fact_schema import resolve_where_bindings
        from semantic_report import parse_semantic_report

        sem = parse_semantic_report(report)
        facts = build_fact_index(report)
        bindings = where if isinstance(where, list) else [where]
        resolved = resolve_where_bindings(bindings, facts, sem)
    except Exception:
        return None
    if not resolved:
        return None
    parts: list[str] = []
    for rf in resolved[:3]:
        label = _node_label(str(rf.name or rf.field or "field"))
        val = rf.value
        field = str(rf.field or "")
        if field == "photo_count":
            parts.append(f"{label}: photos={val}")
        elif field == "caption_count":
            parts.append(f"{label}: captions={val}")
        elif field == "photo_captions":
            caps = [str(c).strip() for c in (val or []) if str(c).strip()] if isinstance(val, list) else []
            parts.append(
                f"{label} captions: {_clip(', '.join(caps), 90)}" if caps else f"{label} captions: (blank)"
            )
        elif field == "result":
            parts.append(f"{label} result={val or '(empty)'}")
        elif field == "attachment_filenames":
            names = [str(n) for n in (val or []) if str(n).strip()] if isinstance(val, list) else []
            parts.append(
                f"{label} files: {_clip(', '.join(names), 90)}" if names else f"{label} files: (none)"
            )
        elif field in ("factory_address",) or "address" in field:
            text = str(val or "").strip()
            parts.append(f"Factory address: {_quote(text, 100)}" if text else "Factory address: (blank)")
        elif field in ("global_remark", "remark") or "remark" in field:
            text = str(val or "").strip()
            parts.append(f"Inspector remark: {_quote(text, 100)}" if text else "Inspector remark: (blank)")
        else:
            if isinstance(val, list):
                text = ", ".join(str(v) for v in val if str(v).strip())
            else:
                text = str(val or "").strip()
            if not text:
                parts.append(f"{label}: (blank)")
            else:
                parts.append(f"{label}: {_quote(text, 100)}")
    return "; ".join(parts) if parts else None


def _defect_rows_snippet(report: dict[str, Any]) -> str | None:
    """Compact live defect list for Original Content (instance photo counts)."""
    try:
        from semantic_report import parse_semantic_report

        sem = parse_semantic_report(report)
        rows: list[str] = []
        for d in sem.defects[:8]:
            name = str(d.get("name") or "?")
            clas = str(d.get("classification") or "?")
            qty = d.get("quantity")
            pc = d.get("photo_count")
            if pc is None:
                photos = d.get("photos")
                pc = len(photos) if isinstance(photos, list) else int(photos or 0)
            rows.append(f"{name} [{clas}] qty={qty} photos={pc}")
    except Exception:
        return None
    if not rows:
        return "Defects: (none recorded)"
    return "Defects: " + "; ".join(rows)


def normalize_node(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (name or "").lower()).strip("_")


def _element_name(el: dict[str, Any]) -> str:
    n = el.get("name") or el.get("title") or ""
    if isinstance(n, dict):
        n = n.get("message") or ""
    return str(n)


def _walk(elements: list[dict[str, Any]]) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    for el in elements:
        found.append(el)
        if el.get("elements"):
            found.extend(_walk(el["elements"]))
        content = (el.get("content") or {}).get("elements")
        if content:
            found.extend(_walk(content))
    return found


def _iter_checklist_elements(report: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for step in (report.get("result") or {}).get("steps", []):
        for action in step.get("actions", []):
            for key in ("testsChecklist", "defectsChecklist"):
                cl = action.get(key) or {}
                out.extend(_walk((cl.get("content") or {}).get("elements", [])))
    return out


def pending_media_kind(spec: dict[str, Any] | None) -> str | None:
    """Return 'photo' | 'attachment' for pending media content — never out_of_report."""
    if not spec:
        return None
    status = str(spec.get("status_class") or "").lower()
    ds = str(spec.get("data_source") or "").lower()
    if status == "advisory" or ds == "out_of_report":
        return None
    processor = str(spec.get("pending_processor") or "").lower()
    bindings = spec.get("where_bindings") or []
    fields: list[str] = []
    for b in bindings:
        if not isinstance(b, dict):
            continue
        if b.get("type") == "out_of_report":
            return None
        field = str(b.get("field") or "").lower()
        sel = str(b.get("selector") or "").lower()
        if field:
            fields.append(field)
        if sel.endswith(".photo_content") or ".photo_content" in sel:
            fields.append("photo_content")
        if sel.endswith(".attachment_content") or ".attachment_content" in sel:
            fields.append("attachment_content")
    if "photo_content" in fields or processor == "vision" or (
        status == "pending" and ds == "report_images"
    ):
        return "photo"
    if "attachment_content" in fields or processor in ("xlsx", "excel", "pdf") or (
        status == "pending" and ds == "report_attachments"
    ):
        return "attachment"
    return None


def _where_from_spec_or_meta(spec: dict[str, Any] | None, meta: dict[str, Any] | None) -> Any:
    if spec and spec.get("where_bindings"):
        # Convert bindings back to authoring-ish where for asset lookup
        out: list[Any] = []
        for b in spec["where_bindings"]:
            if not isinstance(b, dict):
                continue
            if b.get("type") == "selector" and b.get("selector"):
                out.append(str(b["selector"]))
            elif b.get("type") == "intent":
                out.append(
                    {
                        "kind": b.get("kind"),
                        "match": b.get("match") or [],
                        "field": b.get("field") or "comment",
                    }
                )
        return out or (meta or {}).get("where")
    return (meta or {}).get("where")


def _checklist_targets_from_where(where: Any) -> list[str]:
    """Normalized checklist node names / match queries from where."""
    if not where:
        return []
    entries = where if isinstance(where, list) else [where]
    targets: list[str] = []
    for entry in entries:
        if isinstance(entry, str):
            if entry.startswith("checklist."):
                parts = entry.split(".")
                if len(parts) >= 2:
                    targets.append(parts[1])
            continue
        if isinstance(entry, dict):
            if entry.get("kind") in ("checklist", "section") or entry.get("type") == "intent":
                match = entry.get("match") or []
                if isinstance(match, str):
                    match = [match]
                if match:
                    targets.append(" ".join(str(m) for m in match))
                sel = str(entry.get("selector") or "")
                if sel.startswith("checklist."):
                    parts = sel.split(".")
                    if len(parts) >= 2:
                        targets.append(parts[1])
    return targets


def _element_attachment_names(el: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for f in el.get("attachments") or el.get("files") or []:
        if isinstance(f, dict):
            n = str(f.get("name") or f.get("fileName") or f.get("filename") or "").strip()
        else:
            n = str(f).strip()
        if n:
            names.append(n)
    return names


def _element_photo_names(el: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for im in el.get("images") or []:
        if not isinstance(im, dict):
            continue
        n = str(
            im.get("caption")
            or im.get("name")
            or im.get("fileName")
            or im.get("filename")
            or im.get("id")
            or ""
        ).strip()
        if n:
            names.append(n)
    return names


def _where_is_report_attachment(where: Any) -> bool:
    entries = where if isinstance(where, list) else ([where] if where else [])
    for entry in entries:
        if isinstance(entry, str) and entry.startswith("report.") and "attachment" in entry:
            return True
        if isinstance(entry, dict):
            sel = str(entry.get("selector") or "")
            field = str(entry.get("field") or "")
            if sel.startswith("report.") and "attachment" in sel:
                return True
            if entry.get("kind") in ("report", "remark") and "attachment" in field:
                return True
    return False


def _elements_matching_targets(report: dict[str, Any] | None, targets: list[str]) -> list[dict[str, Any]]:
    """Match checklist elements to where targets. No targets → no elements (never all)."""
    if not report or not targets:
        return []
    wanted = [normalize_node(t) for t in targets]
    wanted_compact = [re.sub(r"[^a-z0-9]", "", t.lower()) for t in targets]
    hits: list[dict[str, Any]] = []
    for el in _iter_checklist_elements(report):
        name = _element_name(el)
        n_norm = normalize_node(name)
        n_compact = re.sub(r"[^a-z0-9]", "", name.lower())
        for w, wc in zip(wanted, wanted_compact):
            if not w and not wc:
                continue
            if w and (w == n_norm or w in n_norm or n_norm in w):
                hits.append(el)
                break
            if wc and (wc in n_compact or n_compact in wc):
                hits.append(el)
                break
            # token overlap for intent match queries
            w_tokens = {t for t in w.split("_") if len(t) >= 3}
            n_tokens = {t for t in n_norm.split("_") if len(t) >= 3}
            if w_tokens and len(w_tokens & n_tokens) >= max(1, min(2, len(w_tokens))):
                hits.append(el)
                break
    return hits


def _report_level_attachment_names(report: dict[str, Any] | None) -> list[str]:
    if not report:
        return []
    names: list[str] = []
    for bucket in (
        report.get("attachments") or [],
        (report.get("result") or {}).get("attachments") or [],
    ):
        for name in bucket:
            if isinstance(name, dict):
                n = str(name.get("name") or name.get("fileName") or "").strip()
            else:
                n = str(name).strip()
            if n:
                names.append(n)
    return names


def media_asset_names(
    report: dict[str, Any] | None,
    *,
    kind: str,
    where: Any,
) -> list[str]:
    """File/picture names tied to the pending where binding only."""
    targets = _checklist_targets_from_where(where)
    names: list[str] = []
    if targets:
        for el in _elements_matching_targets(report, targets):
            if kind == "attachment":
                names.extend(_element_attachment_names(el))
            else:
                names.extend(_element_photo_names(el))
    elif kind == "attachment" and _where_is_report_attachment(where):
        names.extend(_report_level_attachment_names(report))
    # Dedupe preserve order
    seen: set[str] = set()
    out: list[str] = []
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def _locus_present_on_report(report: dict[str, Any] | None, *, kind: str, where: Any) -> bool:
    """True only when the bound checklist/report locus exists on this report."""
    targets = _checklist_targets_from_where(where)
    if targets:
        return bool(_elements_matching_targets(report, targets))
    if kind == "attachment" and _where_is_report_attachment(where):
        return report is not None
    return False


def _pending_when_applies(spec: dict[str, Any], report: dict[str, Any] | None) -> bool:
    """Skip pending row when WHEN is known false. Unknown → keep the row."""
    when = spec.get("when")
    if not when:
        return True
    if report is None:
        return True
    try:
        from fact_index import build_fact_index
        from obligation_eval import when_applies_json_only
        from semantic_report import parse_semantic_report

        semantic = parse_semantic_report(report)
        facts = build_fact_index(report)
        applies = when_applies_json_only(spec, facts, semantic=semantic)
        return applies is not False
    except Exception:
        return True


def unable_media_rows_for_order(
    *,
    order_id: str,
    start_index: int,
    specs_by_id: dict[str, dict[str, Any]] | None,
    meta_by_id: dict[str, dict[str, Any]] | None,
    report: dict[str, Any] | None,
    include_order_id: bool,
) -> list[list[Any]]:
    """Manual-check rows for pending photo/attachment CONTENT only (not out_of_report).

    Emits a row only when:
      - spec is pending media content (not advisory/out_of_report)
      - WHEN is not known-false on this report
      - the bound locus exists on this report
    """
    specs_by_id = specs_by_id or {}
    meta_by_id = meta_by_id or {}
    rows: list[list[Any]] = []
    seq = start_index
    for cid, spec in sorted(specs_by_id.items(), key=lambda kv: kv[0]):
        kind = pending_media_kind(spec)
        if not kind:
            continue
        if not _pending_when_applies(spec, report):
            continue
        meta = meta_by_id.get(cid) or {
            "requirement": str(spec.get("requirement") or ""),
            "rule_short": _rule_short_from_requirement(str(spec.get("requirement") or ""), None),
            "field_location": _field_location_from_requirement(
                str(spec.get("requirement") or ""), None
            ),
            "where": None,
        }
        where = _where_from_spec_or_meta(spec, meta)
        if not _locus_present_on_report(report, kind=kind, where=where):
            continue
        assets = media_asset_names(report, kind=kind, where=where)
        locus = str(meta.get("field_location") or checkpoint_column({"checkpoint_id": cid}, meta))
        if kind == "attachment":
            primary = assets[0] if assets else locus
            checkpoint = f"Attachment: {primary}"
            original = ", ".join(assets) if assets else "(no attachment filename found)"
            action = MANUAL_ATTACHMENT_ACTION
        else:
            primary = assets[0] if assets else locus
            checkpoint = f"Photo: {primary}"
            original = ", ".join(assets) if assets else "(no photo name/caption found)"
            action = MANUAL_PHOTO_ACTION
        what_to_check = str(meta.get("rule_short") or meta.get("requirement") or "").strip()
        rows.append(
            [
                order_id if include_order_id and seq == start_index else None,
                None,
                seq,
                WHAT_WAS_CHECKED_UNABLE,
                _clip(checkpoint, 80),
                _clip(original, 140),
                _clip(what_to_check, 160),
                action,
                None,
                None,
            ]
        )
        seq += 1
        include_order_id = False
    return rows


def flag_rows_for_order(
    *,
    order_id: str,
    flags: list[dict[str, Any]],
    requirements: dict[str, str],
    meta_by_id: dict[str, dict[str, Any]] | None = None,
    report: dict[str, Any] | None = None,
    specs_by_id: dict[str, dict[str, Any]] | None = None,
    include_unable_media: bool = False,
) -> list[list[Any]]:
    """Violation rows; optionally pending photo/attachment manual-check rows.

    Pending Unable rows are off by default — they drowned the handoff sheet.
    """
    _ = requirements  # kept for call-site compat; What-was-checked is fixed
    meta_by_id = meta_by_id or {}
    rows: list[list[Any]] = []
    for i, flag in enumerate(flags, start=1):
        cid = str(flag.get("checkpoint_id") or "")
        meta = meta_by_id.get(cid)
        spec = (specs_by_id or {}).get(cid)
        snippet = enrich_report_snippet(
            report,
            str(flag.get("evidence") or ""),
            meta=meta,
            spec=spec,
        )
        # Also try defect/locus enrichment from meta where when atom evidence is junk
        if snippet is None and report is not None:
            where = (meta or {}).get("where") or ""
            if "defect" in str(where).lower() or "Defects" in str((meta or {}).get("field_location") or ""):
                snippet = _defect_rows_snippet(report)
        rows.append(
            [
                order_id if i == 1 else None,
                None,
                i,
                WHAT_WAS_CHECKED,
                checkpoint_column(flag, meta),
                original_content(flag, snippet),
                non_conformity(flag, meta, snippet),
                suggested_actions(flag, meta),
                None,  # Finding Verdict — report team
                None,  # Remark — report team
            ]
        )
    if include_unable_media:
        unable_rows = unable_media_rows_for_order(
            order_id=order_id,
            start_index=len(rows) + 1,
            specs_by_id=specs_by_id,
            meta_by_id=meta_by_id,
            report=report,
            include_order_id=not rows,
        )
        rows.extend(unable_rows)
    return rows


def write_findings_workbook(
    orders: list[tuple[str, list[dict[str, Any]]]],
    requirements: dict[str, str],
    output_path: Path,
    *,
    summary_rows: list[dict[str, Any]] | None = None,
    meta_by_id: dict[str, dict[str, Any]] | None = None,
    reports_by_order: dict[str, dict[str, Any]] | None = None,
    specs_by_id: dict[str, dict[str, Any]] | None = None,
    include_unable_media: bool = False,
) -> Path:
    """Write Findings (+ optional Summary) workbook.

    ``orders`` is a list of ``(order_id, flags)`` in the order they should appear.
    Pending photo/attachment content specs are omitted unless ``include_unable_media``.
    Out-of-report advisory specs are never written.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append(FINDINGS_HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for col, _ in enumerate(FINDINGS_HEADERS, start=1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    reports_by_order = reports_by_order or {}
    for order_id, flags in orders:
        rows = flag_rows_for_order(
            order_id=order_id,
            flags=flags,
            requirements=requirements,
            meta_by_id=meta_by_id,
            report=reports_by_order.get(order_id),
            specs_by_id=specs_by_id,
            include_unable_media=include_unable_media,
        )
        if not rows:
            continue
        for row in rows:
            ws.append(row)

    widths = {
        "A": 14,
        "B": 6,
        "C": 4,
        "D": 28,
        "E": 32,
        "F": 36,
        "G": 52,
        "H": 44,
        "I": 14,
        "J": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = wrap
            if cell.row > 1:
                ws.row_dimensions[cell.row].height = 60
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FINDINGS_HEADERS))}{max(ws.max_row, 1)}"

    if summary_rows:
        sm = wb.create_sheet("Summary")
        keys = list(summary_rows[0].keys())
        sm.append(keys)
        for row in summary_rows:
            sm.append([row.get(k) for k in keys])
        for col in range(1, len(keys) + 1):
            sm.column_dimensions[get_column_letter(col)].width = 18

    legend = wb.create_sheet("Verdict legend")
    legend.append(["Finding Verdict", "Meaning"])
    for verdict, meaning in (
        ("Accurate", "Finding is correct — report should be fixed"),
        ("Extra but ok", "Valid observation but not needed for this handoff"),
        ("Wrong", "False positive — do not send / ignore"),
        ("Pending review", "Not yet judged by the report team"),
    ):
        legend.append([verdict, meaning])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
